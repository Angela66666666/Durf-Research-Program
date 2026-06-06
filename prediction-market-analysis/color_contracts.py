import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 读取数据
ec = pd.read_csv("election_contracts - election_contracts.csv")
sc = pd.read_csv("selected_contracts.csv")

ec_tickers = set(ec["ticker"])
sc_tickers = set(sc["ticker"])

# 颜色定义
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BLUE   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
GREEN  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

wb = Workbook()
ws = wb.active
ws.title = "合约对比"

# 写表头
headers = ec.columns.tolist()
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 黄色：两边都有
row_num = 2
ws.cell(row=row_num, column=1, value="▼ 黄色：两边都有").font = Font(bold=True)
row_num += 1
for _, row in ec[ec["ticker"].isin(sc_tickers)].iterrows():
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=row[h])
        cell.fill = YELLOW
    row_num += 1

# 空一行
row_num += 1

# 蓝色：只在 election_contracts
ws.cell(row=row_num, column=1, value="▼ 蓝色：只在 election_contracts（被删除）").font = Font(bold=True)
row_num += 1
for _, row in ec[~ec["ticker"].isin(sc_tickers)].iterrows():
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=row[h])
        cell.fill = BLUE
    row_num += 1

# 空一行
row_num += 1

# 绿色：只在 selected_contracts
ws.cell(row=row_num, column=1, value="▼ 绿色：只在 selected_contracts（从其他来源加入）").font = Font(bold=True)
row_num += 1
only_in_sc = sc[~sc["ticker"].isin(ec_tickers)]
for _, row in only_in_sc.iterrows():
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=row[h])
        cell.fill = GREEN
    row_num += 1

# 调整列宽
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 50

# 添加图例
legend_row = row_num + 2
ws.cell(row=legend_row,     column=1, value="图例：").font = Font(bold=True)
ws.cell(row=legend_row + 1, column=1, value="黄色").fill = YELLOW
ws.cell(row=legend_row + 1, column=2, value="在 election_contracts 和 selected_contracts 中都有")
ws.cell(row=legend_row + 2, column=1, value="蓝色").fill = BLUE
ws.cell(row=legend_row + 2, column=2, value="在 election_contracts 中有，但 selected_contracts 中没有（被删除）")
ws.cell(row=legend_row + 3, column=1, value="绿色").fill = GREEN
ws.cell(row=legend_row + 3, column=2, value="在 selected_contracts 中有，但 election_contracts 中没有（从其他来源加入）")

output_path = "contracts_comparison.xlsx"
wb.save(output_path)

# 统计
in_both    = len(ec[ec["ticker"].isin(sc_tickers)])
only_in_ec = len(ec[~ec["ticker"].isin(sc_tickers)])
only_in_sc_count = len(only_in_sc)

print(f"完成！文件保存为 {output_path}")
print(f"  黄色（两边都有）：{in_both} 行")
print(f"  蓝色（只在 election_contracts）：{only_in_ec} 行")
print(f"  绿色（只在 selected_contracts）：{only_in_sc_count} 行")
